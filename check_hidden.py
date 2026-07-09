import openpyxl

try:
    wb = openpyxl.load_workbook('シレトコ分析.xlsx')
    print("Sheet names:", wb.sheetnames)
    for name in wb.sheetnames:
        sheet = wb[name]
        print(f"Sheet {name}: state={sheet.sheet_state}")
except Exception as e:
    print(f"Error: {e}")
