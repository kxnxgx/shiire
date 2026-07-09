# -*- coding: utf-8 -*-
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys

def add_sku_comparison_sheet():
    excel_file = 'シレトコ分析.xlsx'
    last_year_csv = '営業日付別売上分析(旧) (1).csv'
    inventory_csv = '在庫一覧 .csv'
    partner_inventory_xlsx = '知床財団在庫表260708.xlsx'
    
    print("データを読み込んでいます...")
    
    # 1. 昨年度の売上データ（CSV）
    if not os.path.exists(last_year_csv):
        print(f"Error: {last_year_csv} が見つかりません。")
        return
    df_last = pd.read_csv(last_year_csv, encoding='cp932')
    df_last_s = df_last[df_last['表記部門名1'] == 'Shiretoko'].copy()
    
    # 2. 今年度の売上データ（Excel）
    if not os.path.exists(excel_file):
        print(f"Error: {excel_file} が見つかりません。")
        return
    try:
        xls = pd.ExcelFile(excel_file)
        target_sheet = '営業日付別売上分析(旧)'
        if target_sheet not in xls.sheet_names and '営業日付別売上分析(旧) (1)' in xls.sheet_names:
            target_sheet = '営業日付別売上分析(旧) (1)'
        df_this = pd.read_excel(excel_file, sheet_name=target_sheet)
        df_this_s = df_this[df_this['表記部門名1'] == 'Shiretoko'].copy()
    except Exception as e:
        print(f"Error: Excelファイルの読み込みに失敗しました。詳細: {e}")
        return

    # 3. 在庫データ（CSV）
    if not os.path.exists(inventory_csv):
        print(f"Error: {inventory_csv} が見つかりません。")
        return
    df_inv = pd.read_csv(inventory_csv, encoding='cp932')
    df_inv_s = df_inv[df_inv['BrandName'] == 'Shiretoko'].copy()
    
    # 品番ごとの売上数量を集計
    last_qty = df_last_s.groupby('3rd Item No.').agg(
        商品名=('商品名', 'first'),
        昨年度数量=('数量', 'sum')
    ).reset_index()
    
    this_qty = df_this_s.groupby('3rd Item No.').agg(
        商品名=('商品名', 'first'),
        今年度数量=('数量', 'sum')
    ).reset_index()
    
    # 在庫数を品番ごとに全店合計
    inv_qty = df_inv_s.groupby('商品コード').agg(
        現在庫数=('現在数量', 'sum')
    ).reset_index().rename(columns={'商品コード': '3rd Item No.'})
    
    # マージして一覧表を作成
    m = pd.merge(last_qty, this_qty, on='3rd Item No.', how='outer', suffixes=('_last', '_this'))
    m = pd.merge(m, inv_qty, on='3rd Item No.', how='outer')
    
    # 新商品の追加
    NEW_PRODUCTS = [
        {'3rd Item No.': 'SNF015001X', '商品名_this': 'SHIRETOKO SCENIC SPOTS Badges/キタキツネ＆ホロベツ'},
        {'3rd Item No.': 'SNF015002X', '商品名_this': 'SHIRETOKO SCENIC SPOTS Badges/ケイマフリ＆フレペの滝'},
        {'3rd Item No.': 'SNF015003X', '商品名_this': 'SHIRETOKO SCENIC SPOTS Badges/エゾシカ＆知床自然センター'},
        {'3rd Item No.': 'SNF015004X', '商品名_this': 'SHIRETOKO SCENIC SPOTS Badges/オジロワシ＆プユニ岬'},
        {'3rd Item No.': 'SNF015005X', '商品名_this': 'SHIRETOKO SCENIC SPOTS Badges/ヒグマ＆開拓小屋コース'},
        {'3rd Item No.': 'SNF016001X', '商品名_this': 'SHIRETOKO SCENIC SPOTS Stickers/キタキツネ＆ホロベツ'},
        {'3rd Item No.': 'SNF016002X', '商品名_this': 'SHIRETOKO SCENIC SPOTS Stickers/ケイマフリ＆フレペの滝'},
        {'3rd Item No.': 'SNF016003X', '商品名_this': 'SHIRETOKO SCENIC SPOTS Stickers/エゾシカ＆知床自然センター'},
        {'3rd Item No.': 'SNF016004X', '商品名_this': 'SHIRETOKO SCENIC SPOTS Stickers/オジロワシ＆プユニ岬'},
        {'3rd Item No.': 'SNF016005X', '商品名_this': 'SHIRETOKO SCENIC SPOTS Stickers/ヒグマ＆開拓小屋コース'},
    ]
    df_new = pd.DataFrame(NEW_PRODUCTS)
    df_new['商品名_last'] = None
    df_new['昨年度数量'] = 0.0
    df_new['今年度数量'] = 0.0
    df_new['現在庫数'] = 0.0
    
    m = pd.concat([m, df_new], ignore_index=True)
    
    # 4. 取引先（知床財団）在庫データ（Excel）のマージ
    partner_stock_col = '取引先在庫数'
    if os.path.exists(partner_inventory_xlsx):
        try:
            sheet_name_decoded = '\u77e5\u5e8a\u8ca1\u56e3\u5728\u5eab\u8868' # '知床財団在庫表'
            df_partner = pd.read_excel(partner_inventory_xlsx, sheet_name=sheet_name_decoded)
            partner_name_map = {
                '知床財団ｵﾘｼﾞﾅﾙｴｺﾊﾞｯｸﾞ2018（森）': 'SNF005X',
                '知床財団ｵﾘｼﾞﾅﾙｴｺﾊﾞｯｸﾞ（海）': 'SNF011X',
                '知床財団ﾛｺﾞｽﾃｯｶｰ': 'SNF014X',
                '知床財団ｴｺﾎﾞﾄﾙ（0.5Lｸﾞﾚｰ）': 'SNF002012X',
                '知床財団ｴｺﾎﾞﾄﾙ（0.5Lﾌﾞﾙｰ）': 'SNF002032X',
                '知床財団ｴｺﾎﾞﾄﾙ（0.5Lｸﾘｱ）': 'SNF002010X',
                '知床財団ｴｺﾎﾞﾄﾙ（0.5Lｼｰﾌｫｰﾑ）': 'SNF002031X',
                '知床財団ｴｺﾎﾞﾄﾙ（1.0Lｸﾞﾚｰ）': 'SNF001012X',
                '知床財団ｴｺﾎﾞﾄﾙ（1.0Lｸﾘｱ）': 'SNF001010X',
                '知床財団ｴｺﾎﾞﾄﾙ（1.0Lﾁｪﾘｰﾌﾞﾛｯｻﾑ）': 'SNF001030X',
                '知床財団ｴｺﾎﾞﾄﾙ（1.0Lﾊﾞﾀｰ）': 'SNF001029X',
                '知床財団ｴｺﾎﾞﾄﾙ（1.0Lｺｯﾄﾝ）': 'SNF001032X',
                '知床財団ｴｺﾎﾞﾄﾙ（1.0Lﾃﾞﾆﾑ）': 'SNF001033X',
                '知床財団ｴｺﾎﾞﾄﾙ（1.0Lｼﾞㄟﾄﾞ）': 'SNF001034X',
                '知床財団ｴｺﾎﾞﾄﾙ（1.0Lｼﾞｪｲﾄﾞ）': 'SNF001034X',
                '知床財団ｴｺﾎﾞﾄﾙ（1.0Lｼｰﾌｫｰﾑ）': 'SNF001011X',
                '知床財団ｴｺ手袋（ﾌｷとｸﾏ）': 'SNF008031X',
                '知床財団ｴｺ軍手（ｳﾆとｺﾝﾌﾞ）': 'SNF008032X',
                '知床財団ロゴ刺繍今治ｶﾞｰｾﾞﾊﾝｶﾁ（ﾈｲﾋﾞｰ)': 'SNF004020X',
                '知床財団ﾛｺﾞ刺繍今治ｶﾞｰｾﾞﾊﾝｶﾁ（ﾍﾞｰｼﾞｭ)': 'SNF004056X',
                '知床財団ﾛｺﾞ刺繍今治ｶﾞｰｾﾞﾊﾝｶﾁ（ｸﾞﾚｰﾍﾞｰｼﾞｭ)': 'SNF004057X',
                '知床財団ﾛｺﾞ刺繍今治ｶﾞｰｾﾞﾊﾝｶﾁ（%e3%83%a2%e3%83%8d%e3%83%94%e3%83%b3%e3%82%af)': 'SNF004058X',
                '知床財団ﾛｺﾞ刺繍今治ｶﾞｰｾﾞﾊﾝｶﾁ（ﾓﾈﾋﾟﾝｸ)': 'SNF004058X',
                '知床財団ﾛｺﾞ刺繍今治ｶﾞｰｾﾞﾊﾝｶﾁ（ﾀ-ｺｲｽﾞﾌﾞﾙｰ)': 'SNF004059X',
                '知床財団ﾛｺﾞ刺繍今治ｶﾞｰｾﾞﾊﾝｶﾁ（ﾀ-...': 'SNF004059X',
                '知床財団ﾛｺﾞ刺繍今治ｶﾞｰｾﾞﾊﾝｶﾁ（ﾀｰｺｲｽﾞﾌﾞﾙｰ)': 'SNF004059X',
                'シレトコ野帳（ヒグマ）': 'SNF009035X',
                'シレトコ野帳（エゾシカ）': 'SNF009036X',
                'シレトコ野帳（エゾクロテン）': 'SNF009037X',
                'シレトコ野帳（エゾリス）': 'SNF009050X',
                'シレトコ野帳（キタキツネ）': 'SNF009051X',
                '知床かばん（フウロ）': 'SNF013002X',
                '知床かばん（リス）': 'SNF013003X',
                '知床かばん（りす）': 'SNF013003X',
            }
            col_name_decoded = '\u5546\u54c1\u540d' # '商品名'
            col_qty_decoded = '\u5728\u5eab\u6570' # '在庫数'
            
            df_partner['3rd Item No.'] = df_partner[col_name_decoded].map(partner_name_map)
            df_partner_sub = df_partner.dropna(subset=['3rd Item No.'])[['3rd Item No.', col_qty_decoded]].rename(columns={col_qty_decoded: partner_stock_col})
            
            m = pd.merge(m, df_partner_sub, on='3rd Item No.', how='left')
            m[partner_stock_col] = m[partner_stock_col].fillna(0).astype(int)
        except Exception as e:
            print(f"Warning: 取引先在庫の読み込みに失敗しました。詳細: {e}")
            m[partner_stock_col] = 0
    else:
        m[partner_stock_col] = 0
    
    # 区分を決定
    def get_category(row):
        has_last = pd.notna(row['商品名_last'])
        has_this = pd.notna(row['商品名_this'])
        if has_last and has_this:
            return '共通取り扱い'
        elif has_last:
            return '昨年度のみ取り扱い'
        else:
            return '今年度のみ取り扱い'
            
    m['区分'] = m.apply(get_category, axis=1)
    
    # 列の整理と補完
    m['商品名'] = m['商品名_this'].fillna(m['商品名_last'])
    # 在庫名からも補完（売上が過去2年なく在庫だけある場合）
    inv_name_map = df_inv_s.set_index('商品コード')['商品名'].to_dict()
    m['商品名'] = m['商品名'].fillna(m['3rd Item No.'].map(inv_name_map))
        
    m['昨年度数量'] = m['昨年度数量'].fillna(0).astype(int)
    m['今年度数量'] = m['今年度数量'].fillna(0).astype(int)
    m['現在庫数'] = m['現在庫数'].fillna(0).astype(int)
    
    # 必要な列だけに整理
    m = m[['3rd Item No.', '商品名', '昨年度数量', '今年度数量', '現在庫数', partner_stock_col, '区分']].rename(
        columns={'3rd Item No.': '商品コード'}
    )
    
    # 新商品の取引先在庫数を 100 に強制設定
    new_sku_mask = m['商品コード'].str.startswith('SNF015') | m['商品コード'].str.startswith('SNF016')
    m.loc[new_sku_mask, partner_stock_col] = 100
    
    # ソート順を整理
    m = m.sort_values(by=['区分', '商品コード'])
    
    print("Excelファイルに「品番数比較」シートを追加・更新しています...")
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        if '品番数比較' in wb.sheetnames:
            del wb['品番数比較']
            
        ws = wb.create_sheet(title='品番数比較')
        
        # タイトル行（1行目）
        ws.merge_cells("A1:D1")
        ws["A1"] = "【Shiretoko】品番数比較・現在庫数一覧"
        ws["A1"].font = Font(name="BIZ UDPゴシック", size=16, bold=True, color="1B5E20")
        ws.row_dimensions[1].height = 30
        
        # 合計行（2行目）
        ws.row_dimensions[2].height = 24
        ws.cell(row=2, column=1, value="合計").font = Font(name="BIZ UDPゴシック", size=11, bold=True)
        ws.cell(row=2, column=1).fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        ws.cell(row=2, column=1).border = Border(top=Side(style='thin', color='1B5E20'), bottom=Side(style='double', color='1B5E20'))
        
        # 見出し（4行目）
        headers = ['商品コード', '商品名', '昨年度数量 (2024/7〜2025/6)', '今年度数量 (2025/7〜2026/6)', '現在庫数', '取引先在庫数', '区分']
        ws.row_dimensions[4].height = 24
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = Font(name="BIZ UDPゴシック", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # データ（5行目〜）
        for row_idx, row_data in enumerate(m.values, 5):
            ws.row_dimensions[row_idx].height = 20
            is_zebra = (row_idx % 2 == 0)
            zebra_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
                top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD')
            )
            
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name="BIZ UDPゴシック", size=10)
                cell.border = thin_border
                if is_zebra:
                    cell.fill = zebra_fill
                    
                # 数値フォーマットと位置
                if col_idx in [3, 4, 5, 6]:
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")
                    
        # 合計行（2行目）の数式設定（データ入力後に行う）
        max_row = ws.max_row
        total_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        total_border = Border(
            top=Side(style='thin', color='1B5E20'), bottom=Side(style='double', color='1B5E20'),
            left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD')
        )
        total_font = Font(name="BIZ UDPゴシック", size=11, bold=True)
        
        for col in range(2, 8):
            col_letter = get_column_letter(col)
            cell = ws.cell(row=2, column=col)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = total_border
            
            # 数量と在庫数の列にSUBTOTALを設定
            if col in [3, 4, 5, 6]:
                cell.value = f"=SUBTOTAL(9, {col_letter}5:{col_letter}{max_row})"
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.value = ""
                
        # オートフィルター
        ws.auto_filter.ref = f"A4:G{max_row}"
        ws.views.sheetView[0].showGridLines = True
        
        # 列幅の自動調整
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row in [1, 2, 3]:
                    continue
                val_str = str(cell.value or '')
                cell_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
                if cell_len > max_len:
                    max_len = cell_len
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        wb.save(excel_file)
        print(f"成功: {excel_file} の「品番数比較」シートを更新しました！")
        
    except PermissionError:
        print(f"Error: {excel_file} が開かれているため書き込みできません。Excelでファイルを閉じてから再実行してください。")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    add_sku_comparison_sheet()
