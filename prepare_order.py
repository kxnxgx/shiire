# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
import argparse
import sys
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 四捨五入関数（非負値用）
def round_half_up(n):
    return int(n + 0.5)

# 店舗名の表記ゆれを統一する関数
def normalize_store_name(name):
    if not isinstance(name, str):
        return name
    name = name.strip()
    # 札幌
    if 'SAPPORO' in name:
        return 'FJALLRAVEN by 3NITY SAPPORO HUTTE'
    # 東京
    if name == 'TOKYO' or '3NITY TOKYO' in name:
        return 'FJALLRAVEN by 3NITY TOKYO'
    # 大阪ルクア
    if 'ルクア大阪' in name:
        return 'FJALLRAVEN by 3NITY ルクア大阪'
    # 心斎橋パルコ / 大丸心斎橋
    if '大丸心斎橋' in name or '心斎橋パルコ' in name:
        return 'FJALLRAVEN by 3NITY 大丸心斎橋'
    # 玉川高島屋 / 小田急町田 (閉店につき実績・在庫を玉川に統合)
    if '玉川高島屋' in name or '小田急町田' in name:
        return 'FJALLRAVEN by 3NITY玉川高島屋S・C'
    return name

def calculate_order(df_combined, brand, logic_type, wholesale_ratio, inventory_file=None, partner_inventory_file=None):
    df_combined = df_combined.copy()
    
    # 廃盤商品の現行商品へのマッピング定義
    SKU_REPLACEMENT_MAP = {
        'SNF00952X': 'SNF009035X',  # 横顔 ヒグマ -> ヒグマ
        'SNF00953X': 'SNF009051X',  # 横顔 キタキツネ -> キタキツネ
        'SNF00954X': 'SNF009036X',  # 横顔 エゾシカ -> エゾシカ
        'SNF00955X': 'SNF009037X',  # 横顔 エゾタヌキ -> エゾクロテン
        'SNF013001X': 'SNF013002X'  # キバナ -> フウロ
    }
    
    PRODUCT_NAME_MAP = {
        'SNF009035X': 'シレトコ野帳/ヒグマ',
        'SNF009051X': 'シレトコ野帳/キタキツネ',
        'SNF009036X': 'シレトコ野帳/エゾシカ',
        'SNF009037X': 'シレトコ野帳/エゾクロテン',
        'SNF013002X': '知床かばん（フウロ）'
    }
    
    # 全体データの商品コードと商品名を置換
    df_combined['3rd Item No.'] = df_combined['3rd Item No.'].replace(SKU_REPLACEMENT_MAP)
    for sku, new_name in PRODUCT_NAME_MAP.items():
        df_combined.loc[df_combined['3rd Item No.'] == sku, '商品名'] = new_name

    # 生データから指定ブランドを抽出
    df = df_combined[df_combined['表記部門名1'] == brand].copy()
    
    if df.empty:
        print(f"Warning: ブランド '{brand}' のデータが見つかりませんでした。")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
        
    # 昨年度と今年度のフラグを作成
    def get_fiscal_year(y, m):
        if m >= 7:
            fy = int(y)
        else:
            fy = int(y) - 1
            
        if fy == 2024:
            return '24年7月～25年1月(7ヶ月実績)'
        elif fy == 2025:
            return '25年7月～26年1月(7ヶ月実績)'
        else:
            return f"{fy}年度(7ヶ月)"
            
    df['FiscalYear'] = df.apply(lambda row: get_fiscal_year(row['year'], row['month']), axis=1)
    
    # 消化期間（7月〜翌年1月の7ヶ月間）に売上データをフィルタリング
    print("消化対象期間（7月〜1月）の売上実績を抽出しています...")
    df = df[df['month'].isin([7, 8, 9, 10, 11, 12, 1])].copy()
    
    # 店舗名の揺れを統一（結合時のブレをなくすため）
    df['店舗名称'] = df['店舗名称'].apply(normalize_store_name)
    
    # 新商品（缶バッジ・ステッカー 各5デザイン）の定義
    NEW_PRODUCTS = [
        {'商品コード': 'SNF015001X', '商品名': 'SHIRETOKO SCENIC SPOTS Badges/キタキツネ＆ホロベツ', '税抜単価': 300, '卸単価': 210},
        {'商品コード': 'SNF015002X', '商品名': 'SHIRETOKO SCENIC SPOTS Badges/ケイマフリ＆フレペの滝', '税抜単価': 300, '卸単価': 210},
        {'商品コード': 'SNF015003X', '商品名': 'SHIRETOKO SCENIC SPOTS Badges/エゾシカ＆知床自然センター', '税抜単価': 300, '卸単価': 210},
        {'商品コード': 'SNF015004X', '商品名': 'SHIRETOKO SCENIC SPOTS Badges/オジロワシ＆プユニ岬', '税抜単価': 300, '卸単価': 210},
        {'商品コード': 'SNF015005X', '商品名': 'SHIRETOKO SCENIC SPOTS Badges/ヒグマ＆開拓小屋コース', '税抜単価': 300, '卸単価': 210},
        {'商品コード': 'SNF016001X', '商品名': 'SHIRETOKO SCENIC SPOTS Stickers/キタキツネ＆ホロベツ', '税抜単価': 300, '卸単価': 210},
        {'商品コード': 'SNF016002X', '商品名': 'SHIRETOKO SCENIC SPOTS Stickers/ケイマフリ＆フレペの滝', '税抜単価': 300, '卸単価': 210},
        {'商品コード': 'SNF016003X', '商品名': 'SHIRETOKO SCENIC SPOTS Stickers/エゾシカ＆知床自然センター', '税抜単価': 300, '卸単価': 210},
        {'商品コード': 'SNF016004X', '商品名': 'SHIRETOKO SCENIC SPOTS Stickers/オジロワシ＆プユニ岬', '税抜単価': 300, '卸単価': 210},
        {'商品コード': 'SNF016005X', '商品名': 'SHIRETOKO SCENIC SPOTS Stickers/ヒグマ＆開拓小屋コース', '税抜単価': 300, '卸単価': 210},
    ]

    # SKUごとの基準単価（最大値）を取得して一元化
    sku_price_map = df_combined.groupby('3rd Item No.')['税抜単価'].max().to_dict()
    for prod in NEW_PRODUCTS:
        sku_price_map[prod['商品コード']] = prod['税抜単価']
    
    # 店舗別・商品別の数量を集計（単価による行分裂を防ぐため、単価は後でマッピングする）
    pivot = df.pivot_table(index=['店舗名称', '3rd Item No.', '商品名'], 
                           columns='FiscalYear', 
                           values='数量', 
                           aggfunc='sum').fillna(0)
                           
    pivot = pivot.reset_index()
    pivot['税抜単価'] = pivot['3rd Item No.'].map(sku_price_map)
    
    col_last_year = '24年7月～25年1月(7ヶ月実績)'
    col_this_year = '25年7月～26年1月(7ヶ月実績)'
    
    if col_last_year not in pivot.columns:
        pivot[col_last_year] = 0
    if col_this_year not in pivot.columns:
        pivot[col_this_year] = 0
        
    # 目標数の算出
    if logic_type == 'average':
        pivot['目標数'] = ((pivot[col_last_year] + pivot[col_this_year]) / 2)
    elif logic_type == 'this_year':
        pivot['目標数'] = pivot[col_this_year]
    elif logic_type == 'last_year':
        pivot['目標数'] = pivot[col_last_year]
    else:
        pivot['目標数'] = 0 
        
    # 売れ筋商品（野帳・エコバッグ）への厚み（1.2倍バッファ）の適用
    # 対象: 野帳 (SNF009...) と エコバッグ (SNF005... および SNF011...)
    def apply_cushion(row):
        sku = str(row['3rd Item No.'])
        target = row['目標数']
        if sku.startswith('SNF009') or sku.startswith('SNF005') or sku.startswith('SNF011'):
            return target * 1.2
        return target
        
    pivot['目標数'] = pivot.apply(apply_cushion, axis=1)
    
    # 目標数を四捨五入して整数にする
    pivot['目標数'] = pivot['目標数'].apply(round_half_up)
        
    # 在庫データの読み込みとマージ
    if inventory_file and os.path.exists(inventory_file):
        print(f"在庫データ {inventory_file} を読み込んでいます...")
        try:
            df_inv = pd.read_csv(inventory_file, encoding='cp932')
            df_inv['拠点名'] = df_inv['拠点名'].apply(normalize_store_name)
            
            df_inv_sub = df_inv[['拠点名', '商品コード', '現在数量']].rename(
                columns={'拠点名': '店舗名称', '商品コード': '3rd Item No.'}
            )
            # 在庫データの商品コードを置換し、重複を合算
            df_inv_sub['3rd Item No.'] = df_inv_sub['3rd Item No.'].replace(SKU_REPLACEMENT_MAP)
            df_inv_sub = df_inv_sub.groupby(['店舗名称', '3rd Item No.'])['現在数量'].sum().reset_index()
            
            pivot = pd.merge(pivot, df_inv_sub, on=['店舗名称', '3rd Item No.'], how='left')
            pivot['現在数量'] = pivot['現在数量'].fillna(0).apply(round_half_up)
            
            # 発注数 = 目標数 - 現在数量 (マイナスの場合は0)
            pivot['発注数'] = np.maximum(0, pivot['目標数'] - pivot['現在数量'])
            
        except Exception as e:
            print(f"在庫データの読み込みエラー: {e}")
            pivot['現在数量'] = 0
            pivot['発注数'] = pivot['目標数']
    else:
        pivot['現在数量'] = 0
        pivot['発注数'] = pivot['目標数']
    
    # 金額の計算
    pivot['発注上代計'] = pivot['発注数'] * pivot['税抜単価']
    pivot['発注下代計'] = (pivot['発注上代計'] * wholesale_ratio).apply(round_half_up)
    
    # 順番を整理
    pivot = pivot[[
        '店舗名称', '3rd Item No.', '商品名', '税抜単価', 
        col_last_year, col_this_year, '目標数', '現在数量', 
        '発注数', '発注上代計', '発注下代計'
    ]].rename(columns={
        col_last_year: '昨年度実績',
        col_this_year: '今年度実績'
    })

    # 新商品の店舗配分
    NEW_PRODUCTS_DISTRIBUTION = {
        'FJALLRAVEN by 3NITY SAPPORO HUTTE': 3,
        'FJALLRAVEN by 3NITY TOKYO': 15,
        'FJALLRAVEN by 3NITY ルクア大阪': 3,
        'FJALLRAVEN by 3NITY 大丸心斎橋': 4,
        'FJALLRAVEN by 3NITY玉川高島屋S・C': 5, # 小田急町田分(2)を加算
    }

    new_rows = []
    for prod in NEW_PRODUCTS:
        for store, target_qty in NEW_PRODUCTS_DISTRIBUTION.items():
            new_rows.append({
                '店舗名称': store,
                '3rd Item No.': prod['商品コード'],
                '商品名': prod['商品名'],
                '税抜単価': prod['税抜単価'],
                '昨年度実績': 0.0,
                '今年度実績': 0.0,
                '目標数': target_qty,
                '現在数量': 0.0,
                '発注数': target_qty,
                '発注上代計': target_qty * prod['税抜単価'],
                '発注下代計': target_qty * prod['卸単価']
            })
    df_new_prod = pd.DataFrame(new_rows)
    pivot = pd.concat([pivot, df_new_prod], ignore_index=True)
    
    # 店舗別の集計サマリー
    store_summary = pivot.groupby('店舗名称').agg(
        目標数=('目標数', 'sum'),
        現在数量=('現在数量', 'sum'),
        発注数=('発注数', 'sum'),
        発注上代計=('発注上代計', 'sum'),
        発注下代計=('発注下代計', 'sum')
    ).reset_index().rename(columns={'現在数量': '現在庫数'})
    
    # --- SKU別集計（取引先在庫情報のマージを含む） ---
    sku_summary = pivot.groupby(['3rd Item No.', '商品名']).agg(
        昨年度実績=('昨年度実績', 'sum'),
        今年度実績=('今年度実績', 'sum'),
        目標数=('目標数', 'sum'),
        店舗在庫数=('現在数量', 'sum'),
        必要発注数=('発注数', 'sum')
    ).reset_index()
    sku_summary['税抜単価'] = sku_summary['3rd Item No.'].map(sku_price_map)
    
    partner_stock_col = '取引先在庫数'
    
    if partner_inventory_file and os.path.exists(partner_inventory_file):
        print(f"取引先在庫データ {partner_inventory_file} を読み込んでいます...")
        try:
            # Unicodeエスケープを用いて安全にシートをロード
            sheet_name_decoded = '\u77e5\u5e8a\u8ca1\u56e3\u5728\u5eab\u8868' # '知床財団在庫表'
            df_partner = pd.read_excel(partner_inventory_file, sheet_name=sheet_name_decoded)
            
            partner_name_map = {
                '知床財団ｵﾘｼﾞﾅﾙｴｺﾊﾞｯｸﾞ2018（森）': 'SNF005X',
                '知床財団ｵﾘｼﾞﾅﾙｴｺﾊﾞｯｸﾞ（海）': 'SNF011X',
                '知床財団ﾛｺﾞｽﾃｯｶｰ': 'SNF014X',
                '知床財団ｴｺﾎﾞﾄﾙ（0.5Lｸﾞﾚｰ）': 'SNF002012X',
                '知床財団ｴｺﾎﾞﾄﾙ（0.5Lﾌﾞﾙｰ）': 'SNF002032X',
                '知床財団ｴｺﾎﾞﾄﾙ（0.5Lｸﾘｱ）': 'SNF002010X',
                '知床財団ｴｺﾎﾞﾄﾙ（0.5Lｼｰﾌｫｰﾑ）': 'SNF002031X',
                '知床財団ｴｺﾎﾞﾄﾙ（1.0Lｸﾞﾚｰ）': 'SNF001012X',
                '知床財団ｴｺﾎﾞﾄﾙ（1.0Lｸﾘｱ）': 'SNF001010X',
                '知床財団ｴｺﾎﾞﾄﾙ（1.0Lﾁｪﾘｰﾌﾞﾛｯｻﾑ）': 'SNF001030X',
                '知床財団ｴｺﾎﾞﾄﾙ（1.0Lﾊﾞﾀｰ）': 'SNF001029X',
                '知床財団ｴｺﾎﾞﾄﾙ（1.0Lｺｯﾄﾝ）': 'SNF001032X',
                '知床財団ｴｺﾎﾞﾄﾙ（1.0Lﾃﾞﾆﾑ）': 'SNF001033X',
                '知床財団ｴｺﾎﾞﾄﾙ（1.0Lｼﾞㄟﾄﾞ）': 'SNF001034X',
                '知床財団ｴｺﾎﾞﾄﾙ（1.0Lｼﾞｪｲﾄﾞ）': 'SNF001034X',
                '知床財団ｴｺﾎﾞﾄﾙ（1.0Lｼｰﾌｫｰﾑ）': 'SNF001011X',
                '知床財団ｴｺ手袋（ﾌｷとｸﾏ）': 'SNF008031X',
                '知床財団ｴｺ軍手（ｳﾆとｺﾝﾌﾞ）': 'SNF008032X',
                '知床財団ロゴ刺繍今治ｶﾞｰｾﾞﾊﾝｶﾁ（ﾈｲﾋﾞｰ)': 'SNF004020X',
                '知床財団ﾛｺﾞ刺繍今治ｶﾞｰｾﾞﾊﾝｶﾁ（ﾍﾞｰｼﾞｭ)': 'SNF004056X',
                '知床財団ﾛｺﾞ刺繍今治ｶﾞｰｾﾞﾊﾝｶﾁ（ｸﾞﾚｰﾍﾞｰｼﾞｭ)': 'SNF004057X',
                '知床財団ﾛｺﾞ刺繍今治ｶﾞｰｾﾞﾊﾝｶﾁ（%e3%83%a2%e3%83%8d%e3%83%94%e3%83%b3%e3%82%af)': 'SNF004058X',
                '知床財団ﾛｺﾞ刺繍今治ｶﾞｰｾﾞﾊﾝｶﾁ（ﾓﾈﾋﾟﾝｸ)': 'SNF004058X',
                '知床財団ﾛｺﾞ刺繍今治ｶﾞｰｾﾞﾊﾝｶﾁ（ﾀ-ｺｲｽﾞﾌﾞﾙｰ)': 'SNF004059X',
                '知床財団ﾛｺﾞ刺繍今治ｶﾞｰｾﾞﾊﾝｶﾁ（ﾀｰｺｲｽﾞﾌﾞﾙｰ)': 'SNF004059X',
                'シレトコ野帳（ヒグマ）': 'SNF009035X',
                'シレトコ野帳（エゾシカ）': 'SNF009036X',
                'シレトコ野帳（エゾクロテン）': 'SNF009037X',
                'シレトコ野帳（エゾリス）': 'SNF009050X',
                'シレトコ野帳（キタキツネ）': 'SNF009051X',
                '知床かばん（フウロ）': 'SNF013002X',
                '知床かばん（リス）': 'SNF013003X',
                '知床かばん（りす）': 'SNF013003X',
            }
            # Unicodeエスケープを用いて安全にカラムを参照
            col_name_decoded = '\u5546\u54c1\u540d' # '商品名'
            col_qty_decoded = '\u5728\u5eab\u6570' # '在庫数'
            
            df_partner['3rd Item No.'] = df_partner[col_name_decoded].map(partner_name_map)
            df_partner_sub = df_partner.dropna(subset=['3rd Item No.'])[['3rd Item No.', col_qty_decoded]].rename(columns={col_qty_decoded: partner_stock_col})
            
            sku_summary = pd.merge(sku_summary, df_partner_sub, on='3rd Item No.', how='left')
            sku_summary[partner_stock_col] = sku_summary[partner_stock_col].fillna(0).astype(int)
        except Exception as e:
            print(f"取引先在庫データの読み込みに失敗しました。詳細: {e}")
            sku_summary[partner_stock_col] = 0
    else:
        sku_summary[partner_stock_col] = 0
            
    # 確定発注数と欠品数
    new_sku_mask = sku_summary['3rd Item No.'].str.startswith('SNF015') | sku_summary['3rd Item No.'].str.startswith('SNF016')
    sku_summary.loc[new_sku_mask, partner_stock_col] = 100

    sku_summary['確定発注数'] = np.minimum(sku_summary['必要発注数'], sku_summary[partner_stock_col])
    sku_summary['取引先欠品数'] = np.maximum(0, sku_summary['必要発注数'] - sku_summary[partner_stock_col])
    sku_summary['確定発注上代計'] = sku_summary['確定発注数'] * sku_summary['税抜単価']
    
    # 新商品の確定発注下代計は卸値(210円)で直接計算する
    def calc_wholesale_amount(row):
        sku = row['3rd Item No.']
        qty = row['確定発注数']
        price = row['税抜単価']
        if str(sku).startswith('SNF015') or str(sku).startswith('SNF016'):
            return round_half_up(qty * 210)
        else:
            return round_half_up(qty * price * wholesale_ratio)

    sku_summary['確定発注下代計'] = sku_summary.apply(calc_wholesale_amount, axis=1)
    
    sku_summary = sku_summary[[
        '3rd Item No.', '商品名', '税抜単価', '昨年度実績', '今年度実績', '目標数', 
        '店舗在庫数', '必要発注数', partner_stock_col, '確定発注数', '取引先欠品数',
        '確定発注上代計', '確定発注下代計'
    ]].rename(columns={'3rd Item No.': '商品コード'})
    
    return pivot, store_summary, sku_summary

def style_sheet(ws, title, sheet_type='detail'):
    # グリッド線の表示を有効化
    ws.views.sheetView[0].showGridLines = True
    
    # 色の定義（フォレストグリーンテーマ）
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid") # 深緑
    total_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")   # 薄緑
    zebra_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")   # 極薄グレー
    
    # フォント定義
    title_font = Font(name="BIZ UDPゴシック", size=16, bold=True, color="1B5E20")
    header_font = Font(name="BIZ UDPゴシック", size=11, bold=True, color="FFFFFF")
    total_font = Font(name="BIZ UDPゴシック", size=11, bold=True, color="000000")
    body_font = Font(name="BIZ UDPゴシック", size=10)
    
    # 罫線定義
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    total_border = Border(
        top=Side(style='thin', color='1B5E20'),
        bottom=Side(style='double', color='1B5E20'),
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD')
    )
    
    # 1. タイトル行（1行目）
    ws.merge_cells("A1:C1")
    ws["A1"] = title
    ws["A1"].font = title_font
    ws.row_dimensions[1].height = 30
    
    max_row = ws.max_row
    max_col = ws.max_column
    
    # 2. 合計行（2行目）の作成とSUBTOTAL関数の設定
    ws.row_dimensions[2].height = 24
    ws.cell(row=2, column=1, value="合計").font = total_font
    ws.cell(row=2, column=1).fill = total_fill
    ws.cell(row=2, column=1).border = total_border
    
    for col in range(2, max_col + 1):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=2, column=col)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = total_border
        
        # 集計種別ごとの合計ロジック
        if sheet_type == 'store_summary':
            cell.value = f"=SUBTOTAL(9, {col_letter}5:{col_letter}{max_row})"
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
        elif sheet_type == 'sku_summary':
            if col >= 4: # 昨年度実績以降
                cell.value = f"=SUBTOTAL(9, {col_letter}5:{col_letter}{max_row})"
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.value = ""
        else: # detail (店舗・商品詳細)
            if col >= 5: # 昨年度実績以降
                cell.value = f"=SUBTOTAL(9, {col_letter}5:{col_letter}{max_row})"
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.value = ""
                
    # 3. 空白行（3行目）
    ws.row_dimensions[3].height = 10
    
    # 4. 見出し行（4行目）のフォーマット
    ws.row_dimensions[4].height = 24
    for col in range(1, max_col + 1):
        cell = ws.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # 5. データ行（5行目以降）のフォーマットとスタイル適用
    for row in range(5, max_row + 1):
        ws.row_dimensions[row].height = 20
        is_zebra = (row % 2 == 0)
        
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = body_font
            cell.border = thin_border
            if is_zebra:
                cell.fill = zebra_fill
                
            if sheet_type == 'store_summary':
                if col > 1:
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")
            elif sheet_type == 'sku_summary':
                if col >= 3:
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")
            else: # detail
                if col >= 4:
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")
                elif col in [1, 2, 3]:
                    cell.alignment = Alignment(horizontal="left")
                    
    ws.auto_filter.ref = f"A4:{get_column_letter(max_col)}{max_row}"
    
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2, 3]:
                continue
            val_str = str(cell.value or '')
            cell_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
            if cell_len > max_len:
                max_len = cell_len
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

def create_presentation_summary_sheet(wb, summary_df, sku_summary_df):
    ws = wb.create_sheet(title='プレゼンサマリー', index=0)
    
    # 基本フォント・スタイル定義
    font_title = Font(name="BIZ UDPゴシック", size=18, bold=True, color="FFFFFF")
    font_subtitle = Font(name="BIZ UDPゴシック", size=12, bold=True, color="1B5E20")
    font_normal = Font(name="BIZ UDPゴシック", size=11)
    font_kpi_val = Font(name="BIZ UDPゴシック", size=24, bold=True)
    font_kpi_alert = Font(name="BIZ UDPゴシック", size=24, bold=True, color="C62828")
    font_kpi_title = Font(name="BIZ UDPゴシック", size=10, color="555555")
    
    fill_header = PatternFill(patternType="solid", fgColor="1B5E20")
    fill_bg = PatternFill(patternType="solid", fgColor="FFFFFF")
    fill_kpi_bg = PatternFill(patternType="solid", fgColor="F0F4F0")
    fill_kpi_alert_bg = PatternFill(patternType="solid", fgColor="FFEBEE")
    fill_table_header = PatternFill(patternType="solid", fgColor="E8F5E9")
    fill_zebra = PatternFill(patternType="solid", fgColor="F9F9F9")
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
                         
    # 全体を白で塗りつぶす
    for r in range(1, 50):
        for c in range(1, 15):
            ws.cell(row=r, column=c).fill = fill_bg
            
    # タイトル
    ws.merge_cells('B2:G3')
    cell = ws.cell(row=2, column=2, value=" 知床オリジナル商品 発注計画 サマリー＆申請ダッシュボード")
    cell.font = font_title
    cell.fill = fill_header
    cell.alignment = Alignment(vertical="center")
    
    # KPI計算
    total_budget = summary_df['発注下代計'].sum() if '発注下代計' in summary_df.columns else 0
    total_qty = summary_df['発注数'].sum() if '発注数' in summary_df.columns else 0
    total_missing = sku_summary_df['取引先欠品数'].sum() if '取引先欠品数' in sku_summary_df.columns else 0
    
    missing_loss = 0
    if '取引先欠品数' in sku_summary_df.columns and '税抜単価' in sku_summary_df.columns:
        missing_rows = sku_summary_df[sku_summary_df['取引先欠品数'] > 0]
        missing_loss = (missing_rows['税抜単価'] * missing_rows['取引先欠品数']).sum()
        
    # KPI 1: 申請予算（下代）
    ws.merge_cells('B5:C5')
    ws.cell(row=5, column=2, value="申請予算 (下代計)").font = font_kpi_title
    ws.merge_cells('B6:C7')
    c_b6 = ws.cell(row=6, column=2, value=f"¥ {int(total_budget):,}")
    c_b6.font = font_kpi_val
    c_b6.alignment = Alignment(horizontal="center", vertical="center")
    for r in range(5, 8):
        for c in range(2, 4):
            ws.cell(row=r, column=c).fill = fill_kpi_bg
            ws.cell(row=r, column=c).border = thin_border
            
    # KPI 2: 確定発注数
    ws.merge_cells('D5:E5')
    ws.cell(row=5, column=4, value="確定発注数").font = font_kpi_title
    ws.merge_cells('D6:E7')
    c_d6 = ws.cell(row=6, column=4, value=f"{int(total_qty):,} 個")
    c_d6.font = font_kpi_val
    c_d6.alignment = Alignment(horizontal="center", vertical="center")
    for r in range(5, 8):
        for c in range(4, 6):
            ws.cell(row=r, column=c).fill = fill_kpi_bg
            ws.cell(row=r, column=c).border = thin_border
            
    # KPI 3: 取引先欠品数 (警告色)
    ws.merge_cells('F5:G5')
    ws.cell(row=5, column=6, value="取引先欠品数 (機会損失)").font = font_kpi_title
    ws.merge_cells('F6:G7')
    c_f6 = ws.cell(row=6, column=6, value=f"{int(total_missing):,} 個")
    c_f6.font = font_kpi_alert
    c_f6.alignment = Alignment(horizontal="center", vertical="center")
    for r in range(5, 8):
        for c in range(6, 8):
            ws.cell(row=r, column=c).fill = fill_kpi_alert_bg
            ws.cell(row=r, column=c).border = thin_border
            
    # 店舗別表
    ws.cell(row=9, column=2, value="■ 店舗別 発注内訳").font = font_subtitle
    ws.cell(row=9, column=4, value="※玉川高島屋には小田急町田の売上・在庫を統合済").font = Font(name="BIZ UDPゴシック", size=9, color="C62828")
    
    store_cols = ['店舗名称', '目標数', '現在庫数', '発注数', '発注下代計']
    
    # ヘッダー
    for i, col_name in enumerate(store_cols):
        c_idx = i + 2
        cell = ws.cell(row=10, column=c_idx, value=col_name)
        cell.font = Font(name="BIZ UDPゴシック", size=10, bold=True)
        cell.fill = fill_table_header
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        
    row_offset = 11
    if all(c in summary_df.columns for c in store_cols):
        for _, row_data in summary_df.iterrows():
            for i, col_name in enumerate(store_cols):
                c_idx = i + 2
                cell = ws.cell(row=row_offset, column=c_idx, value=row_data[col_name])
                cell.font = font_normal
                cell.border = thin_border
                if i > 0:
                    cell.number_format = "#,##0"
                if row_offset % 2 == 0:
                    cell.fill = fill_zebra
            row_offset += 1
            
    # 欠品表
    row_offset += 2
    ws.cell(row=row_offset, column=2, value="■ 取引先欠品リスト（機会損失リスク）").font = font_subtitle
    ws.cell(row=row_offset, column=5, value=f"推定機会損失額(上代): ¥ {int(missing_loss):,}").font = Font(name="BIZ UDPゴシック", size=11, bold=True, color="C62828")
    row_offset += 1
    
    missing_cols = ['商品名', '必要発注数', '取引先欠品数']
    for i, col_name in enumerate(missing_cols):
        c_idx = i + 2
        cell = ws.cell(row=row_offset, column=c_idx, value=col_name)
        cell.font = Font(name="BIZ UDPゴシック", size=10, bold=True)
        cell.fill = fill_table_header
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        
    row_offset += 1
    if '取引先欠品数' in sku_summary_df.columns:
        missing_rows = sku_summary_df[sku_summary_df['取引先欠品数'] > 0]
        for _, row_data in missing_rows.iterrows():
            for i, col_name in enumerate(missing_cols):
                c_idx = i + 2
                cell = ws.cell(row=row_offset, column=c_idx, value=row_data[col_name])
                cell.font = font_normal
                cell.border = thin_border
                if i > 0:
                    cell.number_format = "#,##0"
                if row_offset % 2 == 0:
                    cell.fill = fill_zebra
            row_offset += 1
            
    # 対策テキスト
    row_offset += 2
    ws.cell(row=row_offset, column=2, value="【推奨対策】").font = font_subtitle
    row_offset += 1
    ws.cell(row=row_offset, column=2, value="1. 取引先（知床財団）へ欠品アイテム（特にシレトコ野帳）の早期増産・再生産を要請する。").font = font_normal
    row_offset += 1
    ws.cell(row=row_offset, column=2, value="2. 欠品カラーの代わりに、在庫のある代替カラーの展開を強化する。").font = font_normal
    
    # 算出ロジックの説明
    row_offset += 2
    ws.cell(row=row_offset, column=2, value="■ 発注数の算出ロジック").font = font_subtitle
    row_offset += 1
    ws.cell(row=row_offset, column=2, value="1. 7ヶ月消化モデル: 過去2年間の同期間（7月〜翌年1月）の売上実績を平均してベース目標数を算出。").font = font_normal
    row_offset += 1
    ws.cell(row=row_offset, column=2, value="2. 安全バッファ: 定番の売れ筋商品（野帳・エコバッグ）については、機会損失を防ぐため目標数に1.2倍のバッファを適用。").font = font_normal
    row_offset += 1
    ws.cell(row=row_offset, column=2, value="3. 在庫引き当てとキャップ: 目標数から「店舗現在庫」を引き、さらに「取引先在庫数」を上限（キャップ）として最終発注数を確定。").font = font_normal
    
    # 定番商品への1.2倍バッファ適用内訳
    row_offset += 2
    ws.cell(row=row_offset, column=2, value="■ 定番商品への1.2倍バッファ適用内訳").font = font_subtitle
    row_offset += 1
    
    buffer_headers = ['商品名', '昨期7ヶ月実績', '今期7ヶ月実績', '2年単純平均', 'バッファ後目標数', 'バッファ上乗せ数']
    for i, col_name in enumerate(buffer_headers):
        c_idx = i + 2
        cell = ws.cell(row=row_offset, column=c_idx, value=col_name)
        cell.font = Font(name="BIZ UDPゴシック", size=10, bold=True)
        cell.fill = fill_table_header
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        
    row_offset += 1
    if not sku_summary_df.empty:
        df_buf = sku_summary_df[
            sku_summary_df['商品コード'].astype(str).str.startswith('SNF009') | 
            sku_summary_df['商品コード'].astype(str).str.startswith('SNF005') | 
            sku_summary_df['商品コード'].astype(str).str.startswith('SNF011')
        ].copy()
        
        for _, row_data in df_buf.iterrows():
            last_val = row_data['昨年度実績'] if '昨年度実績' in row_data else 0
            this_val = row_data['今年度実績'] if '今年度実績' in row_data else 0
            target_val = row_data['目標数'] if '目標数' in row_data else 0
            
            avg_val = (last_val + this_val) / 2
            buffer_diff = target_val - avg_val
            
            row_vals = [
                row_data['商品名'],
                last_val,
                this_val,
                avg_val,
                target_val,
                buffer_diff
            ]
            
            for i, val in enumerate(row_vals):
                c_idx = i + 2
                cell = ws.cell(row=row_offset, column=c_idx, value=val)
                cell.font = font_normal
                cell.border = thin_border
                if i > 0:
                    if isinstance(val, float):
                        cell.number_format = "#,##0.0"
                    else:
                        cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")
                if row_offset % 2 == 0:
                    cell.fill = fill_zebra
            row_offset += 1
            
    # 列幅調整
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15

def main():
    parser = argparse.ArgumentParser(description='過去実績と在庫から発注計画を算出する汎用ツール')
    parser.add_argument('--input', type=str, default='シレトコ分析.xlsx', help='入力となる生データのExcelファイル（今年度実績）')
    parser.add_argument('--last_year_csv', type=str, default='営業日付別売上分析(旧) (1).csv', help='昨年度実績のCSVファイル')
    parser.add_argument('--brand', type=str, default='Shiretoko', help='抽出するブランド名（表記部門名1）')
    parser.add_argument('--logic', type=str, choices=['average', 'this_year', 'last_year'], default='average', 
                        help='目標数の計算ロジック（average:平均, this_year:今年度ベース, last_year:昨年度ベース）')
    parser.add_argument('--wholesale_ratio', type=float, default=0.8, help='下代の計算に用いる掛率（デフォルト: 0.8）')
    parser.add_argument('--inventory', type=str, default=None, help='在庫一覧データのCSVファイル')
    parser.add_argument('--partner_inventory', type=str, default='知床財団在庫表260708.xlsx', help='取引先（知床財団）在庫表')
    
    args = parser.parse_args()
    
    print(f"[{args.brand}] の発注計画を計算します...")
    print(f"- 今年度入力ファイル: {args.input}")
    print(f"- 昨年度入力ファイル: {args.last_year_csv}")
    print(f"- 計算ロジック: {args.logic}")
    print(f"- 下代掛率: {args.wholesale_ratio}")
    if args.inventory:
        print(f"- 在庫引当: {args.inventory} を使用")
    if args.partner_inventory:
        print(f"- 取引先在庫: {args.partner_inventory} を使用")
    
    # 1. 今年度データの読み込み（Excel）
    try:
        xls = pd.ExcelFile(args.input)
        target_sheet = '営業日付別売上分析(旧)'
        if target_sheet not in xls.sheet_names and '営業日付別売上分析(旧) (1)' in xls.sheet_names:
            target_sheet = '営業日付別売上分析(旧) (1)'
        df_this = pd.read_excel(args.input, sheet_name=target_sheet)
        
        # year と month カラムを営業日付から確実に再生成
        df_this['営業日付'] = pd.to_datetime(df_this['営業日付'])
        df_this['year'] = df_this['営業日付'].dt.year
        df_this['month'] = df_this['営業日付'].dt.month
    except Exception as e:
        print(f"Error: {args.input} の読み込みに失敗しました。詳細: {e}")
        sys.exit(1)
        
    # 2. 昨年度データの読み込み（CSV）
    df_combined = df_this
    if os.path.exists(args.last_year_csv):
        print(f"昨年度データ {args.last_year_csv} を読み込んでいます...")
        try:
            df_last = pd.read_csv(args.last_year_csv, encoding='cp932')
            # 営業日付をExcelシリアル値からdatetimeに変換
            df_last['営業日付'] = pd.to_datetime(df_last['営業日付'], unit='D', origin='1899-12-30')
            df_last['year'] = df_last['営業日付'].dt.year
            df_last['month'] = df_last['営業日付'].dt.month
            
            # データ結合
            df_combined = pd.concat([df_this, df_last], ignore_index=True)
            print("昨年度と今年度のデータを結合しました。")
        except Exception as e:
            print(f"Warning: 昨年度データの読み込みに失敗しました。今年度データのみで処理します。詳細: {e}")
    else:
        print(f"Warning: {args.last_year_csv} が見つかりません。今年度データのみで処理します。")
        
    detail_df, summary_df, sku_summary_df = calculate_order(
        df_combined, args.brand, args.logic, args.wholesale_ratio, args.inventory, args.partner_inventory
    )
    
    if detail_df.empty:
        print("出力処理をスキップします。")
        sys.exit(0)
        
    output_filename = f"{args.brand}_発注計画.xlsx"
    
    try:
        wb = openpyxl.Workbook()
        
        # 1. 店舗別発注まとめシートの書き込み
        ws_sum = wb.active
        ws_sum.title = '店舗別発注まとめ'
        sum_headers = list(summary_df.columns)
        for col_idx, header in enumerate(sum_headers, 1):
            ws_sum.cell(row=4, column=col_idx, value=header)
        for row_idx, row_data in enumerate(summary_df.values, 5):
            for col_idx, val in enumerate(row_data, 1):
                ws_sum.cell(row=row_idx, column=col_idx, value=val)
                
        # 2. 商品別発注まとめ（取引先在庫対比）シートの書き込み
        ws_sku = wb.create_sheet(title='商品別発注まとめ')
        sku_headers = list(sku_summary_df.columns)
        for col_idx, header in enumerate(sku_headers, 1):
            ws_sku.cell(row=4, column=col_idx, value=header)
        for row_idx, row_data in enumerate(sku_summary_df.values, 5):
            for col_idx, val in enumerate(row_data, 1):
                ws_sku.cell(row=row_idx, column=col_idx, value=val)
                
        # 3. 発注詳細データシートの書き込み
        ws_det = wb.create_sheet(title='発注詳細データ')
        det_headers = list(detail_df.columns)
        for col_idx, header in enumerate(det_headers, 1):
            ws_det.cell(row=4, column=col_idx, value=header)
        for row_idx, row_data in enumerate(detail_df.values, 5):
            for col_idx, val in enumerate(row_data, 1):
                ws_det.cell(row=row_idx, column=col_idx, value=val)
                
        # スタイルの適用
        logic_jp = {'average': '昨年度と今年度の平均', 'this_year': '今年度実績ベース', 'last_year': '昨年度実績ベース'}[args.logic]
        inv_status = "在庫引き当てあり" if args.inventory else "在庫引き当てなし"
        
        style_sheet(ws_sum, f"【{args.brand}】店舗別発注まとめ（7ヶ月消化ベース: {logic_jp}）", sheet_type='store_summary')
        style_sheet(ws_sku, f"【{args.brand}】商品別発注まとめ（7ヶ月消化＋売れ筋1.2倍）", sheet_type='sku_summary')
        style_sheet(ws_det, f"【{args.brand}】発注詳細データ（7ヶ月消化ベース: {logic_jp}）", sheet_type='detail')
        
        # プレゼンサマリーシートの追加
        create_presentation_summary_sheet(wb, summary_df, sku_summary_df)
        
        wb.save(output_filename)
        print(f"成功: {output_filename} を出力しました！")
    except Exception as e:
        print(f"Error: ファイルの出力に失敗しました。詳細: {e}")

if __name__ == "__main__":
    main()
